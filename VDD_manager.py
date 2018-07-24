import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from jira import JIRA
import numpy as np
import VDD_helper as VDD
import os
import time 

Cedar_VDD = "VDD_generator_M185_V5_5_1_Cert_11 July 2018.xlsx"
Toronto_VDD = "VDD_generator_M145_V5_5_1_Delivery_JIRA Systems and Domains 2018-07-12.xlsx"
os.chdir('C:\\Users\\jpgaviri\\iCloudDrive\\Personal\\Python\\VDD_Management')

if __name__== "__main__" :
    Cedar_VDD = VDD.ReadVDD(Cedar_VDD)
    ASI_VDD = VDD.ReadVDD(Toronto_VDD)
    CR_SheetNames = Cedar_VDD.sheetnames
    ASI_SheetNames = ASI_VDD.sheetnames

    TotalDataWS = ASI_VDD['Total Data']
    NumRowsTD = TotalDataWS.max_row
    CR_num = []
    for row in range(2,NumRowsTD+1):
        CR = str(TotalDataWS[row][8].value)
        try:
            CR = CR.split('\n')
        except:
            continue
        CR_num.append(CR)
    #find missing CRs in Total data
    for sheet in CR_SheetNames:
        if (sheet == 'Intro') or (sheet == 'Stats') or (sheet == 'OPRs'):
            continue
        WS = Cedar_VDD[sheet]
        NumRowWS = WS.max_row
        
        for row in range (2,NumRowWS+1):
            CR = str(WS[row][2].value)
            try:
                CR = CR.split('\n')
            except:
                continue
            found = False
            #CR_not found = 'None'
            for line in range(len(CR_num)):
                if found == True:
                    break
                for CRID in CR_num[line]:
                    if found == True:
                        break
                    for WS_CR in CR:
                        if WS_CR == CRID:
                            found = True
                            break        
            if found == False:
                print (CR)
                print ('is not in Total Data')
    # find deleted CRs from total Data
    #found = False
    for line in range(len(CR_num)):
        found = False
        for CRID in CR_num[line]:
            if found == True:
                break
            for sheet in CR_SheetNames:
                if found == True:
                    break
                if (sheet == 'Intro') or (sheet == 'Stats') or (sheet == 'OPRs'):
                    continue
                WS = Cedar_VDD[sheet]
                #print('looking for CR ' + CRID + ' in ' +sheet)
                NumRowWS = WS.max_row
                for row in range (2,NumRowWS+1):
                    if found == True:
                        break
                    CR = str(WS[row][2].value)
                    try:
                        CR = CR.split('\n')
                    except:
                        continue
                    for WS_CR in CR:
                        #print (WS_CR)
                        if WS_CR == CRID:
                            found = True
                            #print ('found')
                            break               
        if found == False:
            print (CRID)
            print ('needs to be deleted from Total Data')
 
    #VDD = Update_TotalData_from_Jira(VDD)
    #VDD = Update_TotalData_from_AllM145(VDD)
    #VDD.save('UpdatedVDD2.xlsx')
