#import xlsxwriter
#import xlrd
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import time
#import csv 

Expenses_File = "expenses.xlsx"
Transactions_File = "transactions_Aug24_2018.xlsx"
os.chdir('C:\\Users\\jpgaviri\\iCloudDrive\\Personal\\Python\\JPGAERO\\Expenses')
def Update_Expenses(Expenses, Transactions):

    

    TransactionsWS = Transactions['transactions_Aug24_2018']
    
    NumRows = TransactionsWS.max_row
    # NumColumn = TransactionsWS.max_column
    # Read FX from Expenses
    ActiveSheet = Expenses['FX']
    Lastrow = ActiveSheet.max_row
    ForeignX = []
    for month in range(2,Lastrow+1):
        ForeignX.append([ActiveSheet.cell(month,1).value,ActiveSheet.cell(month,2).value])

    for transaction in range(2,NumRows):
        date = TransactionsWS.cell(transaction, 1).value.strftime('%d_%m_%Y')
        #print (date)
        # Generate tabs for every month in the expenses file
        match = False
        SheetNames = Expenses.sheetnames
        for sheet in SheetNames:
            if date[3:] == sheet:
                match = True
                break
        if match == False:
            FX = 1.3
            for months in ForeignX:
                if date[3:] == months[0]:
                    FX = months[1]
                    break
            Expenses.create_sheet(date[3:])
            ActiveSheet = Expenses[date[3:]]
            ActiveSheet.cell(1,1).value = date[3:]
            ActiveSheet.cell(1,2).value = "USD Exchange rate"
            ActiveSheet.cell(1,3).value = FX
            ActiveSheet.cell(2,1).value = "Date"
            ActiveSheet.cell(2,2).value = "Description"
            ActiveSheet.cell(2,3).value = "Accounting fees"
            ActiveSheet.cell(2,4).value = "Administrative and General Expenses"
            ActiveSheet.cell(2,5).value = "Advertising"
            ActiveSheet.cell(2,6).value = "Auto Expenses - Distance @ .55"
            ActiveSheet.cell(2,7).value = "Auto Expenses - $ Milleage"
            ActiveSheet.cell(2,8).value = "Auto Expenses - Gas"
            ActiveSheet.cell(2,9).value = "Auto Expenses - Insurance"
            ActiveSheet.cell(2,10).value = "Auto Expenses - Lease"
            ActiveSheet.cell(2,11).value = "Auto Expenses - Maintenance and repairs"
            ActiveSheet.cell(2,12).value = "Auto Expenses - Parking and tolls"
            ActiveSheet.cell(2,13).value = "Bank charges"
            ActiveSheet.cell(2,14).value = "Computer expenses"
            ActiveSheet.cell(2,15).value = "Insurance"
            ActiveSheet.cell(2,16).value = "Interest"
            ActiveSheet.cell(2,17).value = "Internet"
            ActiveSheet.cell(2,18).value = "Licenses, memberships, etc"
            ActiveSheet.cell(2,19).value = "Maintenance"
            ActiveSheet.cell(2,20).value = "Meals and Entertainment"
            ActiveSheet.cell(2,21).value = "Office expenses"
            ActiveSheet.cell(2,22).value = "Other expenses"
            ActiveSheet.cell(2,23).value = "Rent"
            ActiveSheet.cell(2,24).value = "Salaries"
            ActiveSheet.cell(2,25).value = "Subcontracts"
            ActiveSheet.cell(2,26).value = "Supplies"
            ActiveSheet.cell(2,27).value = "Telephone"
            ActiveSheet.cell(2,28).value = "Training"
            ActiveSheet.cell(2,29).value = "Travel Expenses"
            ActiveSheet.cell(2,30).value = "Utilities"
            ActiveSheet.cell(2,31).value = "Income"
            ActiveSheet.cell(2,32).value = "Health"
            ActiveSheet.cell(2,34).value = "Other - Category"
            ActiveSheet.cell(2,35).value = "$ Amount"
            dateraw = TransactionsWS.cell(transaction, 1).value
            LastRow = ActiveSheet.max_row + 1
            DayOfMonth = dateraw.replace(day=1)
            for _ in range(1,31):
                # Charge the rent + parking
                if DayOfMonth.day ==1:
                    ActiveSheet.cell(LastRow,1).value = DayOfMonth 
                    ActiveSheet.cell(LastRow,2).value = "Office Rent"   
                    ActiveSheet.cell(LastRow,23).value = -1133   
                    LastRow +=1
                # on weekdays add the per diem for gas                    
                if DayOfMonth.weekday() != 5 and DayOfMonth.weekday() != 6:
                    ActiveSheet.cell(LastRow,1).value = DayOfMonth
                    ActiveSheet.cell(LastRow,2).value = "Drive to Customer site"   
                    ActiveSheet.cell(LastRow,6).value = 64   
                    ActiveSheet.cell(LastRow,7).value = -35.2
                    LastRow +=1 
                DayOfMonth = DayOfMonth + datetime.timedelta(days=+1)
                if DayOfMonth.day ==1:
                    break
            month = dateraw.month

        # Add each transaction to the correct month
        ActiveSheet = Expenses[date[3:]]
        LastRow = ActiveSheet.max_row
        # Hide transfer transactions as they don't provide much value
        if (TransactionsWS.cell(transaction, 6).value != "Transfer"):
            ActiveSheet.cell(LastRow +1,1).value = TransactionsWS.cell(transaction, 1).value 
            ActiveSheet.cell(LastRow +1,2).value = TransactionsWS.cell(transaction, 2).value
        # if the cost is in a USD card then multiply by the exchange
        if (TransactionsWS.cell(transaction, 7).value == "Starwood Preferred Guest") or \
                (TransactionsWS.cell(transaction, 7).value == "CREDIT CARD") or \
                (TransactionsWS.cell(transaction, 7).value == "Bank of America Cash Rewards Visa Platinum Plus ") or \
                (TransactionsWS.cell(transaction, 7).value == "Green Card"):
            MonthFX =  float(ActiveSheet.cell(1,3).value)
        else :
            MonthFX = 1
        # Multiply debit by -1 and credit by 1
        if (TransactionsWS.cell(transaction, 5).value == "debit"):
            MonthFX *= -1
        # add the cost to each category or to the other at the end
        if (TransactionsWS.cell(transaction, 6).value == "Restaurants") or \
                (TransactionsWS.cell(transaction, 6).value == "Coffee Shops") or \
                (TransactionsWS.cell(transaction, 6).value == "Fast Food") or \
                (TransactionsWS.cell(transaction, 6).value == "Alcohol & Bars") or \
                (TransactionsWS.cell(transaction, 6).value == "Food & Dining"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,20).value = Cost * 0.5
        elif (TransactionsWS.cell(transaction, 6).value == "Accounting Fees"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,3).value = Cost        
        elif (TransactionsWS.cell(transaction, 6).value == "Advertising"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,5).value = Cost        
        elif (TransactionsWS.cell(transaction, 6).value == "Mobile Phone"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,27).value = Cost        
        elif (TransactionsWS.cell(transaction, 6).value == "Internet"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,17).value = Cost        
        elif (TransactionsWS.cell(transaction, 6).value == "Office Supplies"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,26).value = Cost        
        elif (TransactionsWS.cell(transaction, 6).value == "Shipping"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,21).value = Cost        
        elif (TransactionsWS.cell(transaction, 6).value == "Utilities"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,30).value = Cost
        elif (TransactionsWS.cell(transaction, 6).value == "Parking"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,12).value =Cost
        elif (TransactionsWS.cell(transaction, 6).value == "Health & Fitness") or \
                (TransactionsWS.cell(transaction, 6).value == "Gym") or \
                (TransactionsWS.cell(transaction, 6).value == "Doctor"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,32).value = Cost
        elif (TransactionsWS.cell(transaction, 6).value == "Rental Car & Taxi") or \
                (TransactionsWS.cell(transaction, 6).value == "Travel") or \
                (TransactionsWS.cell(transaction, 6).value == "Hotel") or \
                (TransactionsWS.cell(transaction, 6).value == "Air Travel"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,29).value = Cost
        elif (TransactionsWS.cell(transaction, 6).value == "Clothing"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,22).value =Cost
        elif (TransactionsWS.cell(transaction, 6).value == "Paycheck"):
            # little hack to change Salary Juan to negative as on the sheet is possitive
            if MonthFX > 0:
                MonthFX *=-1
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,24).value =Cost
        elif (TransactionsWS.cell(transaction, 6).value == "Income"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,31).value =Cost
        elif (TransactionsWS.cell(transaction, 6).value == "Tuition"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,28).value =Cost
        elif (TransactionsWS.cell(transaction, 6).value == "Home Insurance"):
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,15).value =Cost
        elif (TransactionsWS.cell(transaction, 6).value == "Transfer"):
            continue
        else :
            Cost = float(TransactionsWS.cell(transaction, 4).value * MonthFX)
            ActiveSheet.cell(LastRow +1,34).value = TransactionsWS.cell(transaction, 6).value
            ActiveSheet.cell(LastRow +1,35).value = Cost

    # add the column totals and then the grand total
    SheetNames = Expenses.sheetnames
    for sheet in SheetNames:
        if sheet == 'FX' or sheet == 'Summary':
            continue
        ActiveSheet = Expenses[sheet]
        LastRow = ActiveSheet.max_row
        LastCol = ActiveSheet.max_column
        for j in range(3,LastCol+1):
            if j == 33 or j==6:
                continue
            else:
                RowTotal = 0
                for i in range(3,LastRow):
                    if ActiveSheet.cell(i,j).value == None:
                        continue
                    else:
                        try:
                            RowTotal += float(ActiveSheet.cell(i,j).value)   
                        except:
                            continue  
                ActiveSheet.cell(LastRow +2,j).value = RowTotal
        ActiveSheet.cell(LastRow +2,2).value = "Subtotals"
        ActiveSheet.cell(LastRow +4,12).value = "Total Expenses"
        GrandTotal = 0
        for j in range(3,35):
            if ActiveSheet.cell(LastRow +2,j).value == None:
                continue
            else:
                GrandTotal += ActiveSheet.cell(LastRow +2,j).value         
        ActiveSheet.cell(LastRow +4,14).value = float(GrandTotal) 
    SummarySheet = Expenses['Summary']
    for column in range(4,16):
        month = str(SummarySheet.cell(1,column).value)
        #select sheet with month
        try:
            ActiveSheet = Expenses[month]
        except:
            continue
        Lastrow = ActiveSheet.max_row
        #Income
        SummarySheet.cell(3,column).value = ActiveSheet.cell(Lastrow-2,31).value
        
        #Expenses
        SummarySheet.cell(5,column).value = ActiveSheet.cell(Lastrow-2,3).value
        SummarySheet.cell(6,column).value = ActiveSheet.cell(Lastrow-2,4).value
        SummarySheet.cell(7,column).value = ActiveSheet.cell(Lastrow-2,5).value
        SummarySheet.cell(8,column).value = ActiveSheet.cell(Lastrow-2,7).value
        SummarySheet.cell(9,column).value = ActiveSheet.cell(Lastrow-2,8).value
        SummarySheet.cell(10,column).value = ActiveSheet.cell(Lastrow-2,9).value
        SummarySheet.cell(11,column).value = ActiveSheet.cell(Lastrow-2,10).value
        SummarySheet.cell(12,column).value = ActiveSheet.cell(Lastrow-2,11).value
        SummarySheet.cell(13,column).value = ActiveSheet.cell(Lastrow-2,12).value
        SummarySheet.cell(14,column).value = ActiveSheet.cell(Lastrow-2,13).value
        SummarySheet.cell(15,column).value = ActiveSheet.cell(Lastrow-2,14).value
        SummarySheet.cell(16,column).value = ActiveSheet.cell(Lastrow-2,15).value
        SummarySheet.cell(17,column).value = ActiveSheet.cell(Lastrow-2,16).value
        SummarySheet.cell(18,column).value = ActiveSheet.cell(Lastrow-2,17).value
        SummarySheet.cell(19,column).value = ActiveSheet.cell(Lastrow-2,18).value
        SummarySheet.cell(20,column).value = ActiveSheet.cell(Lastrow-2,19).value
        SummarySheet.cell(21,column).value = ActiveSheet.cell(Lastrow-2,20).value
        SummarySheet.cell(22,column).value = ActiveSheet.cell(Lastrow-2,21).value
        SummarySheet.cell(23,column).value = ActiveSheet.cell(Lastrow-2,22).value
        SummarySheet.cell(24,column).value = ActiveSheet.cell(Lastrow-2,23).value
        SummarySheet.cell(25,column).value = ActiveSheet.cell(Lastrow-2,24).value
        SummarySheet.cell(26,column).value = ActiveSheet.cell(Lastrow-2,25).value
        SummarySheet.cell(27,column).value = ActiveSheet.cell(Lastrow-2,26).value
        SummarySheet.cell(28,column).value = ActiveSheet.cell(Lastrow-2,27).value
        SummarySheet.cell(29,column).value = ActiveSheet.cell(Lastrow-2,28).value
        SummarySheet.cell(30,column).value = ActiveSheet.cell(Lastrow-2,29).value
        SummarySheet.cell(31,column).value = ActiveSheet.cell(Lastrow-2,30).value
        SummarySheet.cell(32,column).value = ActiveSheet.cell(Lastrow-2,32).value
    return Expenses

def ReadFiles():
    wb_Expenses = load_workbook(filename = Expenses_File)
    wb_Transactions = load_workbook(filename = Transactions_File)


    return wb_Expenses, wb_Transactions

if __name__== "__main__" :
    Expenses, Transactions = ReadFiles()
    Expenses = Update_Expenses(Expenses, Transactions)
    now = datetime.datetime.now()
    NewFileName = 'expenses_' + now.strftime("%Y-%m-%d-%H%M") +'.xlsx'
    Expenses.save(NewFileName)
