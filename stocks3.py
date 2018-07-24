# import requests
# from pprint import pprint

# url = 'https://query1.finance.yahoo.com/v10/finance/quoteSummary/AAPL?formatted=true&crumb=ILlIC9tOoXt&lang=en-US&region=US&modules=upgradeDowngradeHistory%2CrecommendationTrend%2CfinancialData%2CearningsHistory%2CearningsTrend%2CindustryTrend%2CindexTrend%2CsectorTrend&corsDomain=finance.yahoo.com'
# r = requests.get(url).json()
# pprint(r)

from bs4 import BeautifulSoup
import urllib3 as url
import certifi as cert
from openpyxl import Workbook
from openpyxl import load_workbook
import time

Index_Name = "NASDAQ"
listOfStocks = []

class Stock():
    def __init__(self,symbol,price,PREV_CLOSE,OPEN,BID,ASK,DAYS_RANGE,\
                 FIFTY_TWO_WK_RANGE,TD_VOLUME,AVERAGE_VOLUME_3MONTH,MARKET_CAP,BETA,\
                 PE_RATIO,EPS_RATIO,EARNINGS_DATE,DIVIDEND_AND_YIELD,EX_DIVIDEND_DATE,ONE_YEAR_TARGET_PRICE):
        self.symbol = symbol
        self.price = price
        self.PREV_CLOSE = PREV_CLOSE
        self.OPEN = OPEN
        self.BID = BID
        self.ASK = ASK
        self.DAYS_RANGE = DAYS_RANGE
        self.FIFTY_TWO_WK_RANGE = FIFTY_TWO_WK_RANGE
        self.TD_VOLUME = TD_VOLUME
        self.AVERAGE_VOLUME_3MONTH = AVERAGE_VOLUME_3MONTH
        self.MARKET_CAP = MARKET_CAP
        self.BETA	 = BETA
        self.PE_RATIO = PE_RATIO
        self.EPS_RATIO = EPS_RATIO
        self.EARNINGS_DATE = EARNINGS_DATE
        self.DIVIDEND_AND_YIELD = DIVIDEND_AND_YIELD
        self.EX_DIVIDEND_DATE = EX_DIVIDEND_DATE
        self.ONE_YEAR_TARGET_PRICE = ONE_YEAR_TARGET_PRICE
        # try:
        #     self.payoutRatio = round((((self.dividendYield/100)*self.price)/self.earningsPerShare)*10000)/100
        # except:
        #     self.payoutRatio = "N/A"
        # self.PERatio = PERatio
    def Display(self):
        print ("")
        print ("Symbol:",self.symbol)
        print ("Price: $",self.price)
        print ("Previous Close: $",self.PREV_CLOSE)
        print ("Open: $",self.OPEN)
        print ("Bid: $",self.BID)
        print ("Ask: $",self.ASK)
        print ("Day's Range: $",self.DAYS_RANGE)
        print ("52 Week Range: $",self.FIFTY_TWO_WK_RANGE)
        print ("Volume: $",self.TD_VOLUME)
        print ("Avg. Volume: $",self.AVERAGE_VOLUME_3MONTH)
        print ("Market Cap: $",self.MARKET_CAP)
        print ("Beta: $",self.BETA)
        print ("PE Ratio (TTM): $",self.PE_RATIO)
        print ("EPS (TTM): $",self.EPS_RATIO)
        print ("Earnings Date: $",self.EARNINGS_DATE)
        print ("Forward Dividend & Yield: $",self.DIVIDEND_AND_YIELD)
        print ("Ex-Dividend Date	: $",self.EX_DIVIDEND_DATE)
        print ("1y Target Est: $",self.ONE_YEAR_TARGET_PRICE)
        print ("Website: https://www.google.com/finance?q=" + self.symbol)

def CreateStockClass(symbol, values):
    try:
        price = (values[0])
    except:
        price = "N/A"
#----------------------
    try:
        PREV_CLOSE = (values[1])
    except:
        PREV_CLOSE = "N/A"
#----------------------
    try:
        OPEN = (values[2])
    except:
        OPEN = "N/A"
#----------------------
    try:
        BID = (values[3])
    except:
        BID = "N/A"
#----------------------
    try:
        ASK = (values[4])
    except:
        ASK = "N/A"
#----------------------
    try:
        DAYS_RANGE = (values[5])
    except:
        DAYS_RANGE = "N/A"
#----------------------
    try:
        FIFTY_TWO_WK_RANGE = (values[6])
    except:
        FIFTY_TWO_WK_RANGE = "N/A"
#----------------------
    try:
        TD_VOLUME = (values[7])
    except:
        TD_VOLUME = "N/A"
#----------------------
    try:
        AVERAGE_VOLUME_3MONTH = (values[8])
    except:
        AVERAGE_VOLUME_3MONTH = "N/A"
#----------------------
    try:
        MARKET_CAP = (values[9])
    except:
        MARKET_CAP = "N/A"
#----------------------
    try:
        BETA = (values[10])
    except:
        BETA = "N/A"
#----------------------
    try:
        PE_RATIO = (values[11])
    except:
        PE_RATIO = "N/A"
#----------------------
    try:
        EPS_RATIO = (values[12])
    except:
        EPS_RATIO = "N/A"
#----------------------
    try:
        EARNINGS_DATE = (values[13])
    except:
        EARNINGS_DATE = "N/A"
#----------------------
    try:
        DIVIDEND_AND_YIELD = (values[14])
    except:
        DIVIDEND_AND_YIELD = "N/A"
#----------------------
    try:
        EX_DIVIDEND_DATE = (values[15])
    except:
        EX_DIVIDEND_DATE = "N/A"
#----------------------
    try:
        ONE_YEAR_TARGET_PRICE = (values[16])
    except:
        ONE_YEAR_TARGET_PRICE = "N/A"
#----------------------


    temp = Stock(symbol,price,PREV_CLOSE,OPEN,BID,ASK,DAYS_RANGE,FIFTY_TWO_WK_RANGE\
                ,TD_VOLUME,AVERAGE_VOLUME_3MONTH,MARKET_CAP,BETA,PE_RATIO\
                ,EPS_RATIO,EARNINGS_DATE,DIVIDEND_AND_YIELD,EX_DIVIDEND_DATE,ONE_YEAR_TARGET_PRICE)
    return temp

def get_stock_data(name):
    listOfValues = []
    http = url.PoolManager(cert_reqs='CERT_REQUIRED', ca_certs=cert.where())
    html_doc = http.request('GET', 'https://finance.yahoo.com/quote/' + name + '?p=' + name)
    print('https://finance.yahoo.com/quote/' + name + '?p=' + name)
    soup = BeautifulSoup(html_doc.data, 'html.parser')
    #print (soup.find("span", class_="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)").get_text())
    listOfValues.append(soup.find("span", class_="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)").get_text())
    #print (soup.find(soup.find("span", class_="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)").get_text()))
    # items = soup.find_all("span", class_="Trsdu(0.3s) ")
    #print(listOfValues[0])
    # for i in items:
    #     #print (i)
    #     item = str(i)
    #     item = item.replace('<span class="Trsdu(0.3s) " data-reactid="','')
    #     item = item.replace('</span>','')
    #     item = item.split('>')[1]
    #     listOfValues.append(item)
    #     print(item)
    items2 = soup.find_all("td",class_="Ta(end) Fw(b) Lh(14px)")
    for i in items2:
        #print(i)
        item = str(i)
        item = item.replace('<span class="Trsdu(0.3s) " data-reactid="','')
        item = item.replace('<td class="Ta(end) Fw(b) Lh(14px)" data-reactid="','')
        item = item.split('data-test="')[1]
        item = item.replace('</span></td>','')
        item = item.replace('</td>','')
        try:
            item = item.split('">')[2]
        except:
            item = item.split('">')[1]

        listOfValues.append(item)
        #print(item)

    #return soup.find("span", class_="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)").get_text()
    #print(listOfValues)
    return listOfValues
def ReadListofStocks(Index_Name):
    ticker = []
    wb = load_workbook(filename = (Index_Name +'.xlsx'))
    IndexWS = wb[Index_Name]

    for i in range(1,IndexWS.max_row):
        ticker.append(str(IndexWS[i][0].value))

    return ticker
if __name__ == "__main__": 
    tickers = ReadListofStocks(Index_Name)
    i = 0
    # ReadListofStocks(Index_Name)
    for t in tickers:
        i += 1
        print ("Downloading... " + t)
        print ("[" + str(i) + "/" + str(len(tickers)) + "]")
        try:
            s = CreateStockClass(t,get_stock_data(t))
            listOfStocks.append(s)
        except:
            continue
        time.sleep(2) 
        #s.Display()
    j = 0
    valuableStocks = []
    for s in listOfStocks:
        if s.price <= 15 and (s.price > s.FIFTY_TWO_WK_RANGE.split(' - ')[1]):
            j += 1
            valuableStocks.append(s.symbol)
            s.Display()
            # webbrowser.open("https://www.google.com/finance?q=" + s.symbol)
        # self.symbol = symbol
        # self.price = price
        # self.PREV_CLOSE = PREV_CLOSE
        # self.OPEN = OPEN
        # self.BID = BID
        # self.ASK = ASK
        # self.DAYS_RANGE = DAYS_RANGE
        # self.FIFTY_TWO_WK_RANGE = FIFTY_TWO_WK_RANGE
        # self.TD_VOLUME = TD_VOLUME
        # self.AVERAGE_VOLUME_3MONTH = AVERAGE_VOLUME_3MONTH
        # self.MARKET_CAP = MARKET_CAP
        # self.BETA	 = BETA
        # self.PE_RATIO = PE_RATIO
        # self.EPS_RATIO = EPS_RATIO
        # self.EARNINGS_DATE = EARNINGS_DATE
        # self.DIVIDEND_AND_YIELD = DIVIDEND_AND_YIELD
        # self.EX_DIVIDEND_DATE = EX_DIVIDEND_DATE
        # self.ONE_YEAR_TARGET_PRICE = ONE_YEAR_TARGET_PRICE

    print ("")
    print (str(j),"results.\n")
    print (valuableStocks)


    