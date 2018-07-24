#import xlsxwriter
#import xlrd
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from jira import JIRA
import numpy as np
import os
import time 

newVDDpath = "VDD 4-18-2018.xlsx"
os.chdir('C:\\Users\\jpgaviri\\iCloudDrive\\Personal\\Python\\VDD_Management')
def Update_TotalData_from_Jira(VDD):

    SheetNames = VDD.sheetnames

    TotalDataWS = VDD['Total Data']
    
    NumRows = TotalDataWS.max_row
    NumColumn = TotalDataWS.max_column
    CRs = ['' for x in range(NumRows+1)]
    newTotalData = VDD.copy_worksheet(TotalDataWS)
    newTotalData.title = 'New TD jira'
    AllM145Issues = VDD.create_sheet('AllM145')
    #for row in range(1,NumRows):
    for row in range(1,NumRows+1):
        CRs[row] = str(TotalDataWS[row][8].value)
    # cleanup list of CRs so that only 1 CR per
    for row in range(1,len(CRs)):
        i=0
        for char in CRs[row]:
            if char == '\n':
                CRs[row] = CRs[row][:i]
            elif char == 'O' and CRs[row][i:i+3] == 'OMS' and i>3:
                CRs[row] = CRs[row][:i]
            elif char == 'I' and CRs[row][i:i+4] == 'IFIS' and i>3:
                CRs[row] = CRs[row][:i]
            elif char == 'E' and CRs[row][i:i+5] == 'EICAS' and i>3:
                CRs[row] = CRs[row][:i]
            elif char == 'F' and CRs[row][i:i+4] == 'FDSA' and i>3:
                CRs[row] = CRs[row][:i]
            i+=1
    options = {'server': 'http://alm.rockwellcollins.com/issues/'}
    jira = JIRA(options, basic_auth=('jpgaviri','PSW*'))#Use RCI password
    # Get all projects viewable by anonymous users.
    projects = jira.projects()
    M145 = jira.project('GLOBALA')
    #M145_issues = jira.search_issues('project=GLOBALA')
    block_size = 50
    block_num = 0
    i = 1
    while True:
        start_idx = block_num*block_size
        M145_issues = jira.search_issues('project=GLOBALA', start_idx, block_size)
        #---------------------

        Wp_type, summary, description, function, fix_version, affected_projects,labels, \
            Operational_impact,Plain_English,Criticality = '','','','','','','','','',''
        for issue in M145_issues:
            print (issue)
            Wp_type, summary, description, function, fix_version, affected_projects,labels, \
                Operational_impact,Plain_English,Criticality = '','','','','','','','','',''
            #english_description = issue.raw['fields']['customfield_20330']
            for field in issue.raw['fields']:
                if field == 'customfield_12909':
                    if issue.raw['fields']['customfield_12909'] != None:
                        Wp_type = issue.raw['fields']['customfield_12909']['value']
                if field == 'summary':
                    if issue.raw['fields']['summary'] != None:
                        summary = issue.raw['fields']['summary']
                if field == 'description':
                    if issue.raw['fields']['description'] != None:
                        description = issue.raw['fields']['description']
                if field == 'customfield_19307':
                    if issue.raw['fields']['customfield_19307'] != None:
                        function = issue.raw['fields']['customfield_19307']['value']
                if field == 'fixVersions':
                    if issue.raw['fields']['fixVersions'] != []:
                        fix_version = issue.raw['fields']['fixVersions'][0]['name']
                if field == 'customfield_13705':
                    if issue.raw['fields']['customfield_13705'] != None:
                        affected_projects = issue.raw['fields']['customfield_13705']
                if field == 'labels':
                    if issue.raw['fields']['labels'] != None:
                        labels = issue.raw['fields']['labels']
                if field == 'customfield_12906':
                    if issue.raw['fields']['customfield_12906'] != None:
                        Operational_impact = issue.raw['fields']['customfield_12906']
                if field == 'customfield_18527':
                    if issue.raw['fields']['customfield_18527'] != None:
                        Plain_English = issue.raw['fields']['customfield_18527']
                if field == 'customfield_13514':
                    if issue.raw['fields']['customfield_13514'] != []:
                        Criticality = issue.raw['fields']['customfield_13514'][0]
            #last_update = issue.raw['fields']['customfield_19063']
            #print ("this seems to have worked")
            label = ''
            if labels != '':
                for l in labels: 
                    label = label + ', ' + str(l)
            label = str(label)
            AllM145Issues.cell(row=i,column=2)
            AllM145Issues.cell(row=i,column=3)
            AllM145Issues[i][0].value = 'CR NUmber: ' + str(issue.key) + '\n' \
                                        'Work Package Type: ' + str(Wp_type) + '\n' \
                                        + 'Fix Version: '+ str(fix_version) +'\n' \
                                        + 'Labels: '+ str(label) + '\n' \
                                        + 'Summary: ' + str(summary) + '\n' \
                                        + 'Operational Impact: '+ str(Operational_impact) +'\n' \
                                        + 'English Description: '+ str(Plain_English) +'\n' \
                                        + 'Description: ' + str(description) +'\n'
            AllM145Issues[i][1].value = str(issue.key)
            AllM145Issues[i][2].value = str(fix_version)
            i+=1
        
        #--------------------
        time.sleep(2) 
        if len(M145_issues) == 0:
            # Retrieve issues until there are no more to come
            break
        block_num += 1

    return VDD

def ReadVDD():
    wb = load_workbook(filename = newVDDpath)
    SheetNames = wb.sheetnames

    return wb

if __name__== "__main__" :
     VDD = ReadVDD()
     VDD = Update_TotalData_from_Jira(VDD)
     VDD.save('UpdatedVDD.xlsx')
