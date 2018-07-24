#import xlsxwriter
#import xlrd
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from jira import JIRA
import numpy as np
import os
import time 

newVDDpath = "UpdatedVDD_06_22_2018.xlsx"
os.chdir('C:\\Users\\jpgaviri\\iCloudDrive\\Personal\\Python\\VDD_Management')
def Update_TotalData_from_AllM145(VDD):

    #SheetNames = VDD.sheetnames

    TotalDataWS = VDD['Total Data']
    NumRowsTD = TotalDataWS.max_row
    NumColumnTD = TotalDataWS.max_column
    CRs = ['' for x in range(NumRowsTD+1)]
    #for row in range(1,NumRows):
    for row in range(1,NumRowsTD+1):
        CRs[row] = str(TotalDataWS[row][8].value)
    # cleanup list of CRs so that only 1 CR per
    for row in range(1,len(CRs)):
        i=0
        for char in CRs[row]:
            if char == '\n':
                CRs[row] = CRs[row][:i]
            elif char == 'O' and CRs[row][i:i+3] == 'OMS' and i>3:
                CRs[row] = CRs[row][:i]
            elif char == 'I' and CRs[row][i:i+4] == 'IFIS' and i>3:
                CRs[row] = CRs[row][:i]
            elif char == 'E' and CRs[row][i:i+5] == 'EICAS' and i>3:
                CRs[row] = CRs[row][:i]
            elif char == 'F' and CRs[row][i:i+4] == 'FDSA' and i>3:
                CRs[row] = CRs[row][:i]
            i+=1
    AllM145 = VDD['AllM145']
    NumRowsAll = AllM145.max_row
    NumColumnAll = AllM145.max_column
    newCR = 0
    for issue in range(1,NumRowsAll):
        #print ("Entering for loop")

        if AllM145[issue][2].value == "M145 5.5.1":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5.1 Doc":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5 Black Label Doc":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5 Black Label":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5.0.1 Doc":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5.0.1":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5.0a":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'                
#---------------------
        if AllM145[issue][2].value == "M145 5.5.0":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5.0 Doc":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5.0 IB4a":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5.0 IB4":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5.0 IB3":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5.0 IB2":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5.0 IB1b":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'
#---------------------
        if AllM145[issue][2].value == "M145 5.5.0 IB1":
            #scan for all globals on TD and see if there is a match
            found = False
            print (AllM145[issue][1].value)
            print ("found 1 match")
            line = 1
            for line in range(1,len(CRs)):
                #print (CRs[row])
                if CRs[line] == AllM145[issue][1].value:
                    found = True
                    print ("CR found in total data")
                    break
            if found ==True:
                print ("jumping to next CR")
                continue
            if found == False:
                print ("Adding CR to the total data")
                newCR +=1
                print (line+newCR)
                TotalDataWS.cell(row=line+newCR,column=1)
                TotalDataWS[line+newCR][0].value = AllM145[issue][0].value
                TotalDataWS[line+newCR][7].value = AllM145[issue][1].value
                TotalDataWS[line+newCR][1].value = 'new CR from Jira'


    # FixVersion = ''
    # for issue in range(len(NumRowsAll)):
    #     for char in AllM145[issue][0]:
    #         if AllM145[issue][char] == 'F' and AllM145[issue][char:char+12] == "Fix Version:":
    #             #Scan for '\n" position
    #             i = char
    #             found = False
    #             while found == False:
    #                 if AllM145[issue][i] == '\n':
    #                     found = True
    #                 else:
    #                     i+=1
    #         FixVersion = AllM145[issue][char:i]
    #         break
            



    return VDD

def ReadVDD():
    wb = load_workbook(filename = newVDDpath)
    SheetNames = wb.sheetnames

    return wb

if __name__== "__main__" :
     VDD = ReadVDD()
     #VDD = Update_TotalData_from_Jira(VDD)
     VDD = Update_TotalData_from_AllM145(VDD)
     VDD.save('UpdatedVDD2.xlsx')
